# -*- coding: utf-8 -*-

# ===================================================================================
# ✅ Windows Logger Lite
#
# Author: Walter
# Copyright: Copyright @ 2025, Walter
# Version: 1.9.0 
#
# Description: A discreet, privacy-first computer monitoring tool
#              that records system activity locally.
# ===================================================================================

import os
import sys
import time
import datetime
import uuid
import json
import logging
import threading
import wmi
import psutil
import winreg
import subprocess
from pathlib import Path
import locale
import msoffcrypto
import ntplib
from datetime import timezone
import win32gui
import win32process
from openpyxl import Workbook
import ctypes
from tkinter import Tk
import re
import email_service

# ===================================================================================
# --- CONFIGURATION & CONSTANTS ---
# ===================================================================================
TASK_NAME = "WindowsLoggerLite_StartupLog"
REG_KEY_PATH = r"SOFTWARE\WindowsLoggerLite"
REG_FLAG_NAME = "Installed"
BASE_DIR_PREF = "D:\\SystemLog"
BASE_DIR_FALLBACK = "C:\\SystemLog"
CACHE_SUBDIR = "cache"
HARDWARE_LOG_DIR = "Hardware"
EVENTS_LOG_DIR = "Events"
EXCEL_PASSWORD = "WindowsLogger"
LHM_DOWNLOAD_URL = "https://github.com/LibreHardwareMonitor/LibreHardwareMonitor/releases/"

# ===================================================================================
# --- GLOBAL INITIALIZATIONS ---
# ===================================================================================
BASE_PATH, CACHE_PATH, COMPUTER_UUID, wmi_con, wmi_lhm, LANG = [None] * 6

try:
    wmi_con = wmi.WMI()
except Exception as e:
    wmi_con = None
    print(f"Critical error during main WMI initialization: {e}")

# ===================================================================================
# ⚙️ HELPER, SETUP & DEPLOYMENT FUNCTIONS
# ===================================================================================
#<editor-fold desc="SETUP & HELPERS">
def get_os_language_code():
    try:
        lang_code = locale.getdefaultlocale()[0].lower()
        if 'zh_cn' in lang_code or 'zh' == lang_code: return 'zh_CN'
        if 'zh_tw' in lang_code or 'zh_hk' in lang_code: return 'zh_TW'
        if 'fr' in lang_code: return 'fr'
        if 'es' in lang_code: return 'es'
        if 'ru' in lang_code: return 'ru'
        if 'ar' in lang_code: return 'ar'
        return 'en'
    except Exception:
        return 'en'

def load_language_data(lang_code):
    try:
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.dirname(__file__)
        
        lang_file = Path(base_path) / "lang" / f"{lang_code}.json"
        
        if not lang_file.exists():
            logging.warning(f"Language file for '{lang_code}' not found, falling back to English.")
            lang_file = Path(base_path) / "lang" / "en.json"

        with open(lang_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"CRITICAL: Failed to load language file. Error: {e}")
        logging.critical(f"Failed to load language file. Error: {e}")
        sys.exit(1) # Exit if language files can't be loaded

def copy_to_clipboard(text):
    try:
        r = Tk()
        r.withdraw()
        r.clipboard_clear()
        r.clipboard_append(text)
        r.update()
        r.destroy()
        logging.info(f"Successfully copied to clipboard: {text}")
    except Exception as e:
        logging.error(f"Failed to copy to clipboard: {e}")

def is_lhm_process_running():
    try:
        for proc in psutil.process_iter(['name']):
            if proc.info['name'].lower() == 'librehardwaremonitor.exe':
                return True
    except (psutil.NoSuchProcess, psutil.AccessDenied):
        pass
    return False

def lhm_checker_and_notifier():
    time.sleep(60)
    try:
        lhm_check_obj = wmi.WMI(namespace="root\\LibreHardwareMonitor")
        if lhm_check_obj.Hardware():
            logging.info("LHM WMI data detected. Notification cancelled.")
            return
    except wmi.x_wmi: pass
    except Exception as e:
        logging.error(f"Error during delayed LHM WMI check: {e}")
        return
    if is_lhm_process_running():
        logging.info("LHM process is running but WMI data is not yet available. Suppressing notification.")
        return

    logging.info("LHM WMI and process not detected. Displaying notification.")
    title = LANG['prompts']['lhm_title']
    text = LANG['prompts']['lhm_text']
    copy_to_clipboard(LHM_DOWNLOAD_URL)
    ctypes.windll.user32.MessageBoxW(0, text, title, 0x40 | 0x1000)

def require_admin():
    try:
        is_admin = ctypes.windll.shell32.IsUserAnAdmin() != 0
    except AttributeError:
        is_admin = os.getuid() == 0
    if not is_admin:
        ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join(sys.argv), None, 1)
        sys.exit(0)

def setup_logger(log_path):
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', filename=log_path, filemode='a')

def create_scheduled_task():
    try:
        program_path = sys.executable
        command = f"""
        $action = New-ScheduledTaskAction -Execute '"{program_path}"'
        $trigger = New-ScheduledTaskTrigger -AtLogOn
        $principal = New-ScheduledTaskPrincipal -GroupId 'BUILTIN\\Administrators' -RunLevel Highest
        $settings = New-ScheduledTaskSettingsSet -AllowStartIfOnBatteries -DontStopIfGoingOnBatteries
        Register-ScheduledTask -TaskName '{TASK_NAME}' -Action $action -Trigger $trigger -Principal $principal -Settings $settings -Force -ErrorAction Stop
        """
        subprocess.run(["powershell", "-NoProfile", "-Command", command], check=True, capture_output=True, text=True, creationflags=subprocess.CREATE_NO_WINDOW)
        logging.info(f"Task '{TASK_NAME}' created/updated successfully.")
        return True
    except subprocess.CalledProcessError as e:
        logging.error(f"Failed to create scheduled task. Stderr: {e.stderr}")
        return False
    except Exception as e:
        logging.error(f"An unexpected error occurred while creating task: {e}")
        return False

def perform_first_run_setup():
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, REG_KEY_PATH, 0, winreg.KEY_READ): return
    except FileNotFoundError:
        print("Performing first-run setup...")
        logging.info("First run detected. Creating scheduled task.")
        if create_scheduled_task():
            try:
                with winreg.CreateKey(winreg.HKEY_CURRENT_USER, REG_KEY_PATH) as key:
                    winreg.SetValueEx(key, REG_FLAG_NAME, 0, winreg.REG_SZ, "1")
                logging.info("Registry flag set successfully.")
                print("Scheduled task created. The program will run automatically on next login.")
            except Exception as e: logging.error(f"Failed to set registry flag: {e}")
        else:
            print("Failed to create scheduled task.")
        sys.exit(0)

def get_computer_uuid():
    return str(uuid.getnode())

def setup_directories():
    global BASE_PATH, CACHE_PATH
    
    def is_drive_removable(drive_letter):
        try:
            for part in psutil.disk_partitions(all=True):
                if part.device.lower().startswith(drive_letter.lower()):
                    return 'removable' in part.opts
        except Exception:
            return False
        return False

    def get_log_path():
        pref_path = Path(BASE_DIR_PREF)
        
        if pref_path.drive and Path(pref_path.drive).exists():
            if is_drive_removable(pref_path.drive):
                title = LANG['prompts']['removable_drive_title']
                text = LANG['prompts']['removable_drive_text']
                
                response = ctypes.windll.user32.MessageBoxW(0, text, title, 0x03 | 0x20)
                
                if response == 6: # Yes
                    return BASE_DIR_PREF
                elif response == 7: # No
                    return BASE_DIR_FALLBACK
                else: # Cancel or closed
                    logging.info("User cancelled operation at removable drive prompt.")
                    sys.exit(0)
            else:
                return BASE_DIR_PREF
        return BASE_DIR_FALLBACK

    try:
        base_dir_to_use = get_log_path()
        BASE_PATH = Path(base_dir_to_use)
        BASE_PATH.mkdir(parents=True, exist_ok=True)
    except (IOError, OSError) as e:
        logging.error(f"Failed to create base directory, falling back to user profile: {e}")
        BASE_PATH = Path(os.path.expanduser("~")) / "WindowsLoggerLite"
        BASE_PATH.mkdir(parents=True, exist_ok=True)

    CACHE_PATH = BASE_PATH / CACHE_SUBDIR
    try:
        for dir_name in [HARDWARE_LOG_DIR, EVENTS_LOG_DIR]:
            (BASE_PATH / dir_name).mkdir(parents=True, exist_ok=True)
            (CACHE_PATH / dir_name).mkdir(parents=True, exist_ok=True)
        return True
    except (IOError, OSError) as e:
        logging.error(f"Failed to create subdirectories: {e}")
        return False
#</editor-fold>

# ===================================================================================
# 🖥️ DATA COLLECTION & FILE HANDLING
# ===================================================================================
#<editor-fold desc="DATA COLLECTION">
def get_ntp_time_offset():
    if get_windows_time_settings()[0] == LANG['status']['enabled']:
        return f"{LANG['status']['unexecuted']} ({LANG['status']['enabled']})", 0.0
    try:
        client = ntplib.NTPClient()
        response = client.request('pool.ntp.org', version=3, timeout=10)
        offset = response.offset
        return LANG['status']['success'], round(offset, 3)
    except Exception as e:
        logging.warning(f"NTP check failed: {e}")
        return f"{LANG['status']['failure']}: {type(e).__name__}", 0.0

def get_windows_time_settings():
    auto_time_enabled, auto_timezone_enabled = LANG['status']['disabled'], LANG['status']['disabled']
    try:
        with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SYSTEM\CurrentControlSet\Services\W32Time\Parameters", 0, winreg.KEY_READ) as key:
            if winreg.QueryValueEx(key, "Type")[0] == "NTP": auto_time_enabled = LANG['status']['enabled']
    except Exception: pass
    try:
        with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SYSTEM\CurrentControlSet\Services\tzautoupdate", 0, winreg.KEY_READ) as key:
            if winreg.QueryValueEx(key, "Start")[0] in [2, 3]: auto_timezone_enabled = LANG['status']['enabled']
    except Exception: pass
    return auto_time_enabled, auto_timezone_enabled

def get_region_info():
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Control Panel\International\Geo", 0, winreg.KEY_READ) as key:
            country_code, _ = winreg.QueryValueEx(key, "Name")
            return LANG['region_map'].get(country_code, country_code)
    except Exception: return "N/A"

def get_timezone_str():
    offset_seconds = -time.timezone if (time.daylight == 0) else -time.altzone
    hours, remainder = divmod(abs(offset_seconds), 3600)
    minutes, _ = divmod(remainder, 60)
    sign = "+" if offset_seconds >= 0 else "-"
    return f"UTC{sign}{int(hours):02d}:{int(minutes):02d}"

def _get_wifi_band_from_api():
    net_band = 'N/A'
    hClient = ctypes.c_void_p()
    pInterfaceList = ctypes.c_void_p()
    try:
        wlanapi = ctypes.windll.LoadLibrary('wlanapi.dll')
        if wlanapi.WlanOpenHandle(2, None, ctypes.byref(ctypes.c_ulong()), ctypes.byref(hClient)) != 0: return net_band
        class GUID(ctypes.Structure): _fields_ = [('Data1', ctypes.c_ulong), ('Data2', ctypes.c_ushort), ('Data3', ctypes.c_ushort), ('Data4', ctypes.c_ubyte * 8)]
        class WLAN_INTERFACE_INFO(ctypes.Structure): _fields_ = [('InterfaceGuid', GUID), ('strInterfaceDescription', ctypes.c_wchar * 256), ('isState', ctypes.c_uint)]
        class WLAN_INTERFACE_INFO_LIST(ctypes.Structure): _fields_ = [('dwNumberOfItems', ctypes.c_ulong), ('dwIndex', ctypes.c_ulong), ('InterfaceInfo', WLAN_INTERFACE_INFO * 1)]
        pInterfaceList = ctypes.POINTER(WLAN_INTERFACE_INFO_LIST)()
        if wlanapi.WlanEnumInterfaces(hClient, None, ctypes.byref(pInterfaceList)) != 0: return net_band
        if pInterfaceList.contents.dwNumberOfItems > 0:
            interface_info = pInterfaceList.contents.InterfaceInfo[0]
            dwChannelSize = ctypes.c_ulong(ctypes.sizeof(ctypes.c_ulong))
            pChannel = ctypes.c_ulong()
            if wlanapi.WlanQueryInterface(hClient, ctypes.byref(interface_info.InterfaceGuid), 10, None, ctypes.byref(dwChannelSize), ctypes.byref(pChannel), None) == 0:
                channel = pChannel.value
                if 1 <= channel <= 14: net_band = '2.4 GHz'
                elif 36 <= channel <= 196: net_band = '5 GHz'
                elif channel > 196: net_band = '6 GHz'
    except Exception as e: logging.warning(f"Could not get WiFi band from API: {e}")
    finally:
        if pInterfaceList and 'wlanapi' in locals(): wlanapi.WlanFreeMemory(pInterfaceList)
        if hClient and 'wlanapi' in locals(): wlanapi.WlanCloseHandle(hClient, None)
    return net_band

def get_wifi_details():
    ssid, net_type, net_band = "N/A", "N/A", "N/A"
    try:
        sys_encoding = locale.getpreferredencoding()
        result_process = subprocess.run(['netsh', 'wlan', 'show', 'interfaces'], encoding=sys_encoding, errors='ignore', capture_output=True, check=False, creationflags=subprocess.CREATE_NO_WINDOW)
        
        if "拒绝访问" in result_process.stdout or "access is denied" in result_process.stdout.lower():
            logging.warning("Failed to get WiFi details: Access Denied. This may be due to Location Services being disabled in Windows Privacy Settings.")
        elif result_process.returncode == 0:
            result = result_process.stdout
            ssid_match = re.search(r"SSID\s+:\s(.*)", result, re.IGNORECASE)
            if ssid_match: ssid = ssid_match.group(1).strip()
            type_match = re.search(r"(802\.11[a-zA-Z]{1,2})", result)
            if type_match: net_type = type_match.group(1)
        
        net_band = _get_wifi_band_from_api()
    except Exception as e: logging.warning(f"Unexpected error getting WiFi details: {e}")
    return ssid, net_type, net_band

def get_mac_address_from_wmi():
    if not wmi_con: return "N/A"
    try:
        for adapter in wmi_con.Win32_NetworkAdapterConfiguration(IPEnabled=True):
            if adapter.MACAddress: return adapter.MACAddress
    except Exception: pass
    return "N/A"

def get_static_computer_info():
    if not wmi_con: return {}
    os_info, cs_info = wmi_con.Win32_OperatingSystem()[0], wmi_con.Win32_ComputerSystem()[0]
    auto_time, auto_tz = get_windows_time_settings()
    ntp_status, time_offset = get_ntp_time_offset()
    mac_address = get_mac_address_from_wmi()
    install_date_str = "N/A"
    if hasattr(os_info, 'InstallDate') and os_info.InstallDate:
        try:
            install_date_obj = datetime.datetime.strptime(os_info.InstallDate.split('.')[0], '%Y%m%d%H%M%S')
            install_date_str = install_date_obj.strftime('%Y-%m-%d %H:%M:%S')
        except Exception: pass
    
    primary_ip = "N/A"
    try:
        primary_adapter = wmi_con.Win32_NetworkAdapterConfiguration(IPEnabled=True)
        if primary_adapter and primary_adapter[0].IPAddress:
            primary_ip = primary_adapter[0].IPAddress[0]
    except Exception: pass

    return {"device_name": cs_info.Name, "processor": [p.Name for p in wmi_con.Win32_Processor()], "gpu": [gpu.Name for gpu in wmi_con.Win32_VideoController()], "ram_manufacturer": [mem.Manufacturer for mem in wmi_con.Win32_PhysicalMemory()], "ram_part_number": [mem.PartNumber.strip() for mem in wmi_con.Win32_PhysicalMemory() if mem.PartNumber] or ["N/A"], "ram_total": round(int(cs_info.TotalPhysicalMemory) / (1024**3), 2), "disk_model": [d.Model for d in wmi_con.Win32_DiskDrive()], "disk_capacity": [round(int(d.Size) / (1024**3), 2) for d in wmi_con.Win32_DiskDrive()], "net_adapter_model": [n.Description for n in wmi_con.Win32_NetworkAdapter() if getattr(n, 'NetConnectionID', None) is not None], "mac_address": mac_address, "ip_address": primary_ip, "timezone": get_timezone_str(), "region": get_region_info(), "auto_time_status": auto_time, "auto_timezone_status": auto_tz, "ntp_status": ntp_status, "time_offset": time_offset, "device_id": wmi_con.Win32_ComputerSystemProduct()[0].UUID, "product_id": os_info.SerialNumber, "windows_version": os_info.Caption, "windows_version_num": getattr(os_info, 'Version', 'N/A'), "install_date": install_date_str, "os_build": f"Build {os_info.BuildNumber}"}

def _get_lhm_sensors_universal():
    global wmi_lhm
    sensors_data = []
    
    if not wmi_lhm:
        try:
            wmi_lhm = wmi.WMI(namespace="root\\LibreHardwareMonitor")
        except wmi.x_wmi: wmi_lhm = None; return []
        except Exception: wmi_lhm = None; return []

    try:
        hardware_devices = wmi_lhm.Hardware()
        if hardware_devices:
            for device in hardware_devices:
                query = f"SELECT * FROM Sensor WHERE Parent = '{device.Identifier}'"
                sensors = wmi_lhm.query(query)
                for sensor in sensors:
                    sensors_data.append({'device_type': device.HardwareType, 'sensor': sensor})
            if sensors_data: return sensors_data
    except wmi.x_wmi_invalid_class: pass
    except Exception as e: logging.warning(f"Error during modern LHM query, will try legacy. Error: {e}")

    try:
        sensors = wmi_lhm.Sensor()
        if sensors:
            for sensor in sensors:
                sensors_data.append({'device_type': 'Unknown', 'sensor': sensor})
    except Exception as e: logging.warning(f"Error during legacy LHM query: {e}")
        
    return sensors_data

last_disk_io, last_net_io, last_io_time = psutil.disk_io_counters(perdisk=True), psutil.net_io_counters(pernic=True), time.time()

def get_hardware_snapshot():
    global last_disk_io, last_net_io, last_io_time, wmi_lhm

    current_time = time.time()
    time_delta = current_time - last_io_time
    last_io_time = current_time
    if time_delta <= 0: time_delta = 1
    
    cpu_temp, fan_speeds, gpu_temps, gpu_loads, disk_temps = "N/A", [], [], [], []
    
    try:
        all_sensors = _get_lhm_sensors_universal()
        
        cpu_package_temps = []
        for item in all_sensors:
            device_type = item['device_type']
            sensor = item['sensor']
            s_name_lower = sensor.Name.lower()
            
            if device_type == 'Cpu' and sensor.SensorType == 'Temperature':
                if 'core average' in s_name_lower:
                    cpu_temp = round(sensor.Value, 2)
                elif 'package' in s_name_lower or 'tctl/tdie' in s_name_lower:
                    cpu_package_temps.append(sensor.Value)
            
            elif 'Gpu' in device_type and sensor.SensorType == 'Temperature':
                gpu_temps.append(round(sensor.Value, 2))
            elif 'Gpu' in device_type and sensor.SensorType == 'Load' and 'core' in s_name_lower:
                gpu_loads.append(round(sensor.Value, 2))
            elif 'Storage' in device_type and sensor.SensorType == 'Temperature':
                disk_temps.append(round(sensor.Value, 2))
            elif sensor.SensorType == 'Fan':
                fan_speeds.append(int(sensor.Value))

        if cpu_temp == "N/A" and cpu_package_temps:
            cpu_temp = round(max(cpu_package_temps), 2)

    except wmi.x_wmi as e:
        logging.warning(f"Connection to LHM lost ({e}). Will retry next cycle.")
        wmi_lhm = None
    except Exception as e:
        logging.error(f"Critical error processing LHM sensors: {e}", exc_info=True)

    mem = psutil.virtual_memory()
    current_disk_io = psutil.disk_io_counters(perdisk=True)
    disk_read_speeds, disk_write_speeds, disk_avail_spaces = [], [], []
    for disk_name, start_io in last_disk_io.items():
        end_io = current_disk_io.get(disk_name)
        if end_io:
            disk_read_speeds.append(round((end_io.read_bytes - start_io.read_bytes) / (1024**2) / time_delta, 3))
            disk_write_speeds.append(round((end_io.write_bytes - start_io.write_bytes) / (1024**2) / time_delta, 3))
    for part in psutil.disk_partitions(all=False):
        try:
            if 'fixed' in part.opts.lower() or 'nvme' in part.fstype.lower(): disk_avail_spaces.append(round(psutil.disk_usage(part.mountpoint).free / (1024**3), 2))
        except Exception: continue
    
    ssid, net_type, net_band = get_wifi_details()
    
    current_net_io = psutil.net_io_counters(pernic=True)
    net_adapters, net_upload_speeds, net_download_speeds = [], [], []
    for adapter_name, io_counters in current_net_io.items():
        last_io = last_net_io.get(adapter_name)
        if last_io and (io_counters.bytes_sent > last_io.bytes_sent or io_counters.bytes_recv > last_io.bytes_recv):
            net_adapters.append(adapter_name)
            net_upload_speeds.append(round(((io_counters.bytes_sent - last_io.bytes_sent) * 8 / (1024**2)) / time_delta, 3))
            net_download_speeds.append(round(((io_counters.bytes_recv - last_io.bytes_recv) * 8 / (1024**2)) / time_delta, 3))
    
    last_disk_io, last_net_io = current_disk_io, current_net_io
    gpu_count = len(wmi_con.Win32_VideoController()) if wmi_con else 1
    disk_count = len(disk_read_speeds) if disk_read_speeds else 1
    
    return {"timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "cpu_util": round(psutil.cpu_percent(interval=None), 2), "cpu_temp": cpu_temp, "fan_speed": fan_speeds or ["N/A"], "mem_util": round(mem.percent, 2), "mem_avail": round(mem.available / (1024**3), 2), "gpu_util": gpu_loads or ["N/A"] * gpu_count, "gpu_temp": gpu_temps or ["N/A"] * gpu_count, "disk_read": disk_read_speeds or [0.0] * disk_count, "disk_write": disk_write_speeds or [0.0] * disk_count, "disk_avail": disk_avail_spaces or ["N/A"] * disk_count, "disk_temp": disk_temps or ["N/A"] * disk_count, "net_adapter": net_adapters or ["N/A"], "net_ssid": ssid, "net_type": net_type, "net_band": net_band, "net_upload": net_upload_speeds or [0.0], "net_download": net_download_speeds or [0.0],}
#</editor-fold>

#<editor-fold desc="FILE HANDLING & REPORTING">
def cache_data(data, log_type):
    try:
        ts = datetime.datetime.now().strftime("%Y%m%d%H%M%S_%f")
        cache_dir = CACHE_PATH / (log_type.capitalize())
        with open(cache_dir / f"{ts}.json", 'w', encoding='utf-8') as f: json.dump(data, f, ensure_ascii=False)
    except Exception as e: logging.error(f"Failed to cache data for {log_type}: {e}")

def _create_single_report(date_str, data_type, data_list):
    if not data_list:
        logging.info(f"No '{data_type}' data cached for {date_str}, skipping Excel report.")
        return True
    try:
        info_data = get_static_computer_info()
        wb = Workbook()
        wb.remove(wb.active)
        
        ws = wb.create_sheet(title=LANG['logs']['sheets'][data_type])
        
        list_keys = [k for k, v in data_list[0].items() if isinstance(v, list)]
        max_list_cols = {key: max((len(row.get(key, [])) for row in data_list), default=0) for key in list_keys}
        
        header_row_keys = LANG['logs']['columns'][data_type].keys()
        header_row_display = []
        for key in header_row_keys:
            val = LANG['logs']['columns'][data_type][key]
            num_items = max_list_cols.get(key, 1)
            if num_items > 1:
                for i in range(num_items): header_row_display.append(f"{val} #{i+1}")
            else: header_row_display.append(val)
        ws.append(header_row_display)

        for row_data in data_list:
            row_to_write = []
            for key in header_row_keys:
                val = row_data.get(key)
                if data_type == 'events' and key == 'event_type':
                    val = LANG['logs']['event_types'].get(val, val)
                if key in max_list_cols:
                    padded_val = (val or []) + ["N/A"] * (max_list_cols.get(key, 0) - len(val or []))
                    row_to_write.extend(padded_val)
                else: row_to_write.append(val)
            ws.append(row_to_write)
        
        ws_info = wb.create_sheet(title=LANG['logs']['sheets']['info'])
        for key, header in LANG['logs']['columns']['info'].items():
            value = info_data.get(key, "N/A")
            if isinstance(value, list):
                if value:
                    for i, item in enumerate(value): ws_info.append([f"{header} #{i+1}", item])
                else: ws_info.append([header, "N/A"])
            else: ws_info.append([header, value])

        full_tz = get_timezone_str()
        tz_match = re.search(r"UTC([+-])(\d{2}):\d{2}", full_tz)
        tz_string = f"UTC{tz_match.group(1)}{int(tz_match.group(2))}" if tz_match else "UTC"
        
        file_suffix = LANG['logs']['file_suffixes'][data_type]
        output_dir = BASE_PATH / (HARDWARE_LOG_DIR if data_type == 'hardware' else EVENTS_LOG_DIR)
        
        temp_dir = CACHE_PATH / "temp"
        temp_dir.mkdir(exist_ok=True)
        unencrypted_filename = temp_dir / f"tmp_{uuid.uuid4()}.xlsx"
        final_filename = output_dir / f"{COMPUTER_UUID}_{date_str}_{tz_string}_{file_suffix}.xlsx"
        
        wb.save(unencrypted_filename)
        with open(unencrypted_filename, "rb") as f_in, open(final_filename, "wb") as f_out:
            office_file = msoffcrypto.OfficeFile(f_in)
            office_file.encrypt(EXCEL_PASSWORD, f_out)
        os.remove(unencrypted_filename)
        os.chmod(final_filename, 0o444)
        logging.info(f"Successfully created encrypted report: {final_filename}")

    except Exception as e:
        logging.error(f"Failed to create '{data_type}' report for {date_str}: {e}", exc_info=True)
        return False
    return True

def process_cached_data():
    today_str = datetime.date.today().strftime("%Y-%m-%d")
    files_by_day = {}
    all_files_to_delete = {}
    for log_type in ['hardware', 'events']:
        cache_dir = CACHE_PATH / (log_type.capitalize())
        for f in sorted(cache_dir.glob("*.json")):
            try:
                file_date_str = f.name.split('_')[0][:8]
                file_date = datetime.datetime.strptime(file_date_str, "%Y%m%d").date()
                date_str = file_date.strftime("%Y-%m-%d")
                if date_str != today_str:
                    if date_str not in files_by_day:
                        files_by_day[date_str] = {'hardware': [], 'events': []}
                        all_files_to_delete[date_str] = []
                    with open(f, 'r', encoding='utf-8') as jf:
                        files_by_day[date_str][log_type].append(json.load(jf))
                    all_files_to_delete[date_str].append(f)
            except Exception as e: logging.warning(f"Skipping corrupted cache file {f}: {e}")
    for date_str, daily_data in files_by_day.items():
        daily_data['hardware'].sort(key=lambda x: x.get('timestamp', ''))
        daily_data['events'].sort(key=lambda x: x.get('timestamp', ''))
        hardware_success = _create_single_report(date_str, 'hardware', daily_data['hardware'])
        events_success = _create_single_report(date_str, 'events', daily_data['events'])
        if hardware_success and events_success:
             for f in all_files_to_delete.get(date_str, []):
                try: f.unlink()
                except OSError as e: logging.error(f"Failed to delete cache file {f}: {e}")
#</editor-fold>

# ===================================================================================
# ⚙️ BACKGROUND MONITORING THREAD & MAIN EXECUTION
# ===================================================================================
#<editor-fold desc="ProcessMonitor Thread & Main Execution">
class ProcessMonitor(threading.Thread):
    def __init__(self, stop_event):
        super().__init__(daemon=True)
        self.stop_event = stop_event
        self.system_root = os.environ.get("SystemRoot", "C:\\Windows").lower()
        self.seen_pids = {p.pid for p in psutil.process_iter(['pid'])}
        self.logged_apps = {}
    def _is_gui_app(self, pid):
        try:
            hwnd_list = []
            def callback(hwnd, lst):
                if win32gui.IsWindowVisible(hwnd) and win32gui.GetWindowText(hwnd):
                    _, found_pid = win32process.GetWindowThreadProcessId(hwnd)
                    if found_pid == pid:
                        lst.append(hwnd)
                return True
            win32gui.EnumWindows(callback, hwnd_list)
            return len(hwnd_list) > 0
        except Exception: return False
    def _is_user_app_by_path(self, exe_path):
        if not exe_path: return False
        return not exe_path.lower().startswith(self.system_root)
    def run(self):
        while not self.stop_event.is_set():
            try:
                current_pids = {p.pid for p in psutil.process_iter(['pid'])}
                new_pids, dead_pids = current_pids - self.seen_pids, self.seen_pids - current_pids
                for pid in new_pids:
                    try:
                        p = psutil.Process(pid)
                        exe_path = p.exe()
                        if self._is_user_app_by_path(exe_path) or self._is_gui_app(pid):
                            app_name = p.name()
                            self.logged_apps[pid] = (app_name, exe_path)
                            cache_data({"timestamp": datetime.datetime.now().strftime("%H:%M:%S"), "event_type": "start", "app_name": app_name, "path": exe_path}, "events")
                    except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess): continue
                for pid in dead_pids:
                    if pid in self.logged_apps:
                        app_name, exe_path = self.logged_apps.pop(pid)
                        cache_data({"timestamp": datetime.datetime.now().strftime("%H:%M:%S"), "event_type": "close", "app_name": app_name, "path": exe_path}, "events")
                self.seen_pids = current_pids
            except Exception as e: 
                logging.error(f"Error in ProcessMonitor loop: {e}", exc_info=True)
            self.stop_event.wait(5)

def main():
    global LANG, COMPUTER_UUID
    
    require_admin()
    
    LANG = load_language_data(get_os_language_code())
    
    if not setup_directories(): sys.exit(1)
    
    perform_first_run_setup()
    
    setup_logger(BASE_PATH / "error.log")
    
    COMPUTER_UUID = get_computer_uuid()
    lhm_notifier_thread = threading.Thread(target=lhm_checker_and_notifier, daemon=True)
    lhm_notifier_thread.start()
    try: 
        process_cached_data()
    except Exception as e: 
        logging.error(f"Unhandled error during initial cached data processing: {e}", exc_info=True)
    stop_event = threading.Event()
    process_monitor_thread = ProcessMonitor(stop_event)
    process_monitor_thread.start()
    last_day_checked = datetime.date.today()
    try:
        while True:
            cache_data(get_hardware_snapshot(), 'hardware')
            current_day = datetime.date.today()
            if current_day != last_day_checked:
                process_cached_data()
                last_day_checked = current_day
            now = datetime.datetime.now()
            sleep_duration = 60 - now.second - (now.microsecond / 1_000_000.0)
            time.sleep(max(0, sleep_duration))
    except KeyboardInterrupt:
        print("Shutdown signal received.")
    except Exception as e:
        logging.critical(f"A critical error occurred in the main loop: {e}", exc_info=True)
    finally:
        stop_event.set()
        if process_monitor_thread.is_alive(): 
            process_monitor_thread.join()
        print("Logger stopped.")

if __name__ == "__main__":
    main()

#</editor-fold>
